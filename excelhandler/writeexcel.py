import openpyxl
import os


class WriteExcel():
    def __init__(self, filepath):
        self.filepath = filepath

    def __del__(self):
        if self.workbook != None:
            self.workbook.save(self.filepath)
            self.workbook.close()

    def create_workbook(self):
        if os.path.isfile(self.filepath):
            os.remove(self.filepath)

        wb = openpyxl.Workbook()
        wb.save(self.filepath)
        wb.close()

    def open_workbook(self):
        if os.path.isfile(self.filepath) == False:
            self.create_workbook()

        self.workbook = openpyxl.load_workbook(self.filepath)

    def create_sheet(self, sheet_name: str):
        sheet_names = self.workbook.get_sheet_names()
        sheet_count = len(sheet_names)

        sheet_pos = -1
        for index in range(sheet_count):
            if sheet_name == sheet_names[index]:
                sheet_pos = index
                break

        if sheet_pos != -1:
            tmp_sheet = self.workbook.get_sheet_by_name(sheet_name)
            self.workbook.remove_sheet(tmp_sheet)
            sheet_count = sheet_count - 1

        self.sheet = self.workbook.create_sheet(sheet_name, sheet_count)

        if sheet_pos != -1:
            if sheet_pos < sheet_count:
                self.workbook.move_sheet(self.sheet, sheet_pos)

    def set_cell_value(self, row, col, value):
        self.sheet.cell(row, col, value)

    # def set_cell_value(self, cell_name, value):
    #     self.sheet[cell_name] = value

    def get_cell_value(self, row, col):
        cell = openpyxl.Cell(self.sheet, row, col)
        return cell.value

    def get_cells_values(self):
        """
        一次设置多个单元格的值
        :return:
        """
        pass

    def save(self, filepath=None):
        if filepath == None:
            self.workbook.save(self.filepath)
        else:
            self.workbook.save(filepath)

        self.workbook.close()
        self.workbook = None
